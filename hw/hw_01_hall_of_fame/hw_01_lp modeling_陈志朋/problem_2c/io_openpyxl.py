"""
Reading from and writing to an Excel file using openpyxl.
"""

from openpyxl import load_workbook

import constant


def readDataOpenpyxl() -> tuple[dict, dict, dict]:
    """
    Given an Excel file, read production planning data from Excel file using openpyxl.

    Returns
    -------
    productProfits : dict
        The profit per unit for each product
        - Keys: product names (`Product1`, `Product2`)
        - Values: profit per unit
    componentsProductNeeded : nested dict
        The components required to produce one unit of each product at each plant
        - Keys: plant names (`Frame_parts`, `Electrical_components`)
        - Values: dict
            - Keys: product names (`Product1` and `Product2`)
            - Values: components required per unit
    componentsAvailable : dict
        The available components at each plant
        - Keys: plant names (`Frame_parts`, `Electrical_components`)
        - Values: available components
    """

    # Load a workbook from DATA_PATH
    inputBook = load_workbook(constant.DATA_PATH)

    # Find the sheet in which you want to write the solution
    inputSheet = inputBook[constant.SHEET_NAME]

    # Read data from the inputSheet and create dictionaries
    productProfits = {}
    for productName in constant.PRODUCT_NAMES:
        colIndex = constant.INPUT_PROFIT_START_COL + constant.PRODUCT_NAMES.index(productName)
        profitValue = inputSheet.cell(constant.INPUT_PROFIT_START_ROW, colIndex).value
        productProfits[productName] = profitValue

    componentsAvailable = {}
    for i, plantName in enumerate(constant.COMPONENTS_NAMES):
        rowIndex = constant.INPUT_AVAILABLE_START_ROW + i
        colIndex = constant.INPUT_AVAILABLE_START_COL
        hours = inputSheet.cell(rowIndex, colIndex).value
        componentsAvailable[plantName] = hours

    componentsProductNeeded = {}
    for i, plantName in enumerate(constant.COMPONENTS_NAMES):
        rowIndex = constant.INPUT_START_ROW + i
        hoursPerProduct = {}
        for productName in constant.PRODUCT_NAMES:
            colIndex = constant.INPUT_START_COL + constant.PRODUCT_NAMES.index(productName)
            hours = inputSheet.cell(rowIndex, colIndex).value
            hoursPerProduct[productName] = hours
        componentsProductNeeded[plantName] = hoursPerProduct

    return productProfits, componentsProductNeeded, componentsAvailable


def writeDataOpenpyxl(soln, objVal) -> None:
    """
    Write the solution back to the Excel file using openpyxl.

    Parameters
    ----------
    soln : dict
        The optimal solutions
        - Keys: variable names
        - Values: optimal decision variable values
    objVal : float
        The optimal objective function value
    """
    # Load a workbook from DATA_PATH
    outputBook = load_workbook(constant.DATA_PATH)

    # Find the sheet to write the solution
    outputSheet = outputBook[constant.SHEET_NAME]

    # Write the optimal solutions to the outputSheet
    for j, product in enumerate(constant.PRODUCT_NAMES):
        rowIndex = constant.OUTPUT_BATCHES_PRODUCED_START_ROW
        colIndex = constant.OUTPUT_BATCHES_PRODUCED_START_COL + j
        outputSheet.cell(rowIndex, colIndex, soln[product])

    outputSheet.cell(constant.OUTPUT_PROFIT_START_ROW, constant.OUTPUT_PROFIT_START_COL, objVal)

    # Save the workbook
    outputBook.save(constant.DATA_PATH)

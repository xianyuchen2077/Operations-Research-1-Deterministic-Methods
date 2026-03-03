"""
Reading from and writing to an Excel file using openpyxl.
"""

from openpyxl import load_workbook

import constant


def readDataOpenpyxl() -> tuple[dict, dict, dict]:
    """
    Given an Excel file, read production planning data from Excel file using openpyxl.

    Returns
    -------
    productProfits : dict
        The profit per batch for each product
        - Keys: product names (`Product 1`, `Product 2`)
        - Values: profit per unit
    companyProductMaterials : nested dict
        The materials required to produce one unit of each product
        - Keys: material names (`Frame parts`, `Electrical components`)
        - Values: dict
            - Keys: product names (`Product 1`, `Product 2`)
            - Values: materials required
    companyAvailableMaterials : dict
        The available materials on the company
        - Keys: material names (`Frame parts`, `Electrical components`)
        - Values: available materials

    """

    # Load a workbook from DATA_PATH
    inputBook = load_workbook(constant.DATA_PATH)

    # Find the sheet in which you want to write the solution
    inputSheet = inputBook[constant.SHEET_NAME]

    # Read data from the inputSheet and create dictionaries
    productProfits = {}
    for productName in constant.PRODUCT_NAMES:
        colIndex = constant.INPUT_PROFIT_START_COL + constant.PRODUCT_NAMES.index(productName)
        profitValue = inputSheet.cell(constant.INPUT_PROFIT_START_ROW, colIndex).value
        productProfits[productName] = profitValue

    companyAvailableMaterials = {}
    for i, materialName in enumerate(constant.MATERIAL_NAMES):
        rowIndex = constant.INPUT_MATERIALS_AVAILABLE_START_ROW + i
        colIndex = constant.INPUT_MATERIALS_AVAILABLE_START_COL
        materials = inputSheet.cell(rowIndex, colIndex).value
        companyAvailableMaterials[materialName] = materials

    companyProductMaterials = {}
    for i, materialName in enumerate(constant.MATERIAL_NAMES):
        rowIndex = constant.INPUT_MATERIALS_START_ROW + i
        materialsPerProduct = {}
        for productName in constant.PRODUCT_NAMES:
            colIndex = constant.INPUT_MATERIALS_START_COL + constant.PRODUCT_NAMES.index(productName)
            materials = inputSheet.cell(rowIndex, colIndex).value
            materialsPerProduct[productName] =  materials
        companyProductMaterials[materialName] = materialsPerProduct

    return productProfits, companyProductMaterials,  companyAvailableMaterials


def writeDataOpenpyxl(soln, objVal) -> None:
    """
    Write the solution back to the Excel file using openpyxl.

    Parameters
    ----------
    soln : dict
        The optimal solutions
        - Keys: variable names
        - Values: optimal decision variable values
    objVal : float
        The optimal objective function value
    """
    # Load a workbook from DATA_PATH
    outputBook = load_workbook(constant.DATA_PATH)

    # Find the sheet to write the solution
    outputSheet = outputBook[constant.SHEET_NAME]

    # Write the optimal solutions to the outputSheet
    for j, product in enumerate(constant.PRODUCT_NAMES):
        rowIndex = constant.OUTPUT_UNITS_PRODUCED_START_ROW
        colIndex = constant.OUTPUT_UNITS_PRODUCED_START_COL + j
        outputSheet.cell(rowIndex, colIndex, soln[product])

    outputSheet.cell(constant.OUTPUT_PROFIT_START_ROW, constant.OUTPUT_PROFIT_START_COL, objVal)

    # Save the workbook
    outputBook.save(constant.DATA_PATH)

# -*- coding: utf-8 -*-

# io_openpyxl.py
import openpyxl
import constants

def read_data_openpyxl():
    """
    使用openpyxl从Excel文件读取数据
    """
    try:
        workbook = openpyxl.load_workbook(constants.DATA_PATH)
        sheet = workbook[constants.INPUT_SHEET_NAME]
        
        # 读取利润系数
        profit_coef = [
            sheet.cell(row=constants.PROFIT_ROW, column=constants.PROFIT_COL1).value,
            sheet.cell(row=constants.PROFIT_ROW, column=constants.PROFIT_COL2).value
        ]
        
        # 读取资源消耗数据
        frame_usage = [
            sheet.cell(row=constants.FRAME_ROW, column=constants.FRAME_COL1).value,
            sheet.cell(row=constants.FRAME_ROW, column=constants.FRAME_COL2).value
        ]
        
        electronic_usage = [
            sheet.cell(row=constants.ELECTRONIC_ROW, column=constants.ELECTRONIC_COL1).value,
            sheet.cell(row=constants.ELECTRONIC_ROW, column=constants.ELECTRONIC_COL2).value
        ]
        
        resource_usage = [frame_usage, electronic_usage]
        
        # 读取资源可用量
        resource_avail = [
            sheet.cell(row=constants.FRAME_AVAILABLE_ROW, column=constants.FRAME_AVAILABLE_COL).value,
            sheet.cell(row=constants.ELECTRONIC_AVAILABLE_ROW, column=constants.ELECTRONIC_AVAILABLE_COL).value
        ]
        
        # 读取产品2上限
        max_product2 = sheet.cell(row=constants.Max_Product2_ROW, column=constants.Max_Product2_COL).value
        
        workbook.close()
        
        data = {
            "profit": profit_coef,
            "resource_usage": resource_usage,
            "resource_avail": resource_avail,
            "max_product2": max_product2,
            "product_names": constants.PRODUCT_NAMES
        }
        
        print("Openpyxl数据读取成功!")
        return data
        
    except Exception as e:
        print(f"Openpyxl数据读取错误: {e}")
        return None



def write_solution_openpyxl(solution_dict):
    """
    使用openpyxl将求解结果写入Excel的新工作表
    """
    try:
        workbook = openpyxl.load_workbook(constants.DATA_PATH)
        input_sheet = workbook[constants.INPUT_SHEET_NAME]
        
        # 如果输出工作表已存在，则删除它
        if constants.OUTPUT_SHEET_NAME in workbook.sheetnames:
            del workbook[constants.OUTPUT_SHEET_NAME]
        
        # 创建新工作表
        output_sheet = workbook.create_sheet(constants.OUTPUT_SHEET_NAME)
        
        # 复制原工作表的所有内容和格式
        for row in input_sheet.iter_rows():
            for cell in row:
                new_cell = output_sheet.cell(
                    row=cell.row, 
                    column=cell.column, 
                    value=cell.value
                )
                # 复制单元格样式（如果存在）
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection.copy()
                    new_cell.alignment = cell.alignment.copy()
        
        # 更新求解结果
        # 产品1产量
        output_sheet.cell(
            row=constants.OUTPUT_PRODUCT_ROW, 
            column=constants.OUTPUT_PRODUCT_COL1, 
            value=solution_dict['product1']
        )
        
        # 产品2产量
        output_sheet.cell(
            row=constants.OUTPUT_PRODUCT_ROW, 
            column=constants.OUTPUT_PRODUCT_COL2, 
            value=solution_dict['product2']
        )
        
        # 总利润
        output_sheet.cell(
            row=constants.OUTPUT_PROFIT_ROW, 
            column=constants.OUTPUT_PROFIT_COL, 
            value=solution_dict['optimal_profit']
        )
        
        '''
        # 在工作表顶部添加标识
        output_sheet.insert_rows(1)
        output_sheet['A1'] = "Gurobi Solution - " + constants.MODEL_NAME
        output_sheet['A1'].font = openpyxl.styles.Font(bold=True, color="0066CC")
        '''
        
        workbook.save(constants.DATA_PATH)
        workbook.close()
        print("Openpyxl结果写入成功!")
        return True
        
    except Exception as e:
        print(f"Openpyxl结果写入错误: {e}")
        return False
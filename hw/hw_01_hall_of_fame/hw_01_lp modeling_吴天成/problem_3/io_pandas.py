# -*- coding: utf-8 -*-

# io_pandas.py
import pandas as pd
import openpyxl
import constants

def read_data_pandas():
    """
    使用pandas从Excel文件读取数据
    """
    try:
        # 读取整个Data工作表
        df = pd.read_excel(constants.DATA_PATH, sheet_name=constants.INPUT_SHEET_NAME, header=None)
        
        # 读取利润系数
        profit_coef = [
            df.iloc[constants.PROFIT_ROW-1, constants.PROFIT_COL1-1],
            df.iloc[constants.PROFIT_ROW-1, constants.PROFIT_COL2-1]
        ]
        
        # 读取资源消耗数据
        frame_usage = [
            df.iloc[constants.FRAME_ROW-1, constants.FRAME_COL1-1],
            df.iloc[constants.FRAME_ROW-1, constants.FRAME_COL2-1]
        ]
        
        electronic_usage = [
            df.iloc[constants.ELECTRONIC_ROW-1, constants.ELECTRONIC_COL1-1],
            df.iloc[constants.ELECTRONIC_ROW-1, constants.ELECTRONIC_COL2-1]
        ]
        
        resource_usage = [frame_usage, electronic_usage]
        
        # 读取资源可用量
        resource_avail = [
            df.iloc[constants.FRAME_AVAILABLE_ROW-1, constants.FRAME_AVAILABLE_COL-1],
            df.iloc[constants.ELECTRONIC_AVAILABLE_ROW-1, constants.ELECTRONIC_AVAILABLE_COL-1]
        ]
        
        # 读取产品2上限
        max_product2 = df.iloc[constants.Max_Product2_ROW-1, constants.Max_Product2_COL-1]
        
        data = {
            "profit": profit_coef,
            "resource_usage": resource_usage,
            "resource_avail": resource_avail,
            "max_product2": max_product2,
            "product_names": constants.PRODUCT_NAMES
        }
        
        print("Pandas数据读取成功!")
        return data
        
    except Exception as e:
        print(f"Pandas数据读取错误: {e}")
        return None

def write_solution_pandas(solution_dict):
    """
    使用pandas将求解结果写入Excel的新工作表
    """
    try:
        # 使用openpyxl来复制格式，pandas用于数据读取
        workbook = openpyxl.load_workbook(constants.DATA_PATH)
        input_sheet = workbook[constants.INPUT_SHEET_NAME]
        
        # 如果输出工作表已存在，则删除它
        if constants.OUTPUT_SHEET_NAME in workbook.sheetnames:
            del workbook[constants.OUTPUT_SHEET_NAME]
        
        # 创建新工作表
        output_sheet = workbook.create_sheet(constants.OUTPUT_SHEET_NAME)
        
        # 复制原工作表的所有内容和格式
        for row in input_sheet.iter_rows():
            for cell in row:
                new_cell = output_sheet.cell(
                    row=cell.row, 
                    column=cell.column, 
                    value=cell.value
                )
                # 复制单元格样式
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection.copy()
                    new_cell.alignment = cell.alignment.copy()
        
        # 更新求解结果
        output_sheet.cell(
            row=constants.OUTPUT_PRODUCT_ROW, 
            column=constants.OUTPUT_PRODUCT_COL1, 
            value=solution_dict['product1']
        )
        
        output_sheet.cell(
            row=constants.OUTPUT_PRODUCT_ROW, 
            column=constants.OUTPUT_PRODUCT_COL2, 
            value=solution_dict['product2']
        )
        
        output_sheet.cell(
            row=constants.OUTPUT_PROFIT_ROW, 
            column=constants.OUTPUT_PROFIT_COL, 
            value=solution_dict['optimal_profit']
        )
        
        '''
        # 添加标识
        output_sheet.insert_rows(1)
        output_sheet['A1'] = "Gurobi Solution - " + constants.MODEL_NAME
        output_sheet['A1'].font = openpyxl.styles.Font(bold=True, color="0066CC")
        '''
        
        workbook.save(constants.DATA_PATH)
        workbook.close()
        print("Pandas结果写入成功! 已完全复制原表格格式。")
        return True
        
    except Exception as e:
        print(f"Pandas结果写入错误: {e}")
        return False
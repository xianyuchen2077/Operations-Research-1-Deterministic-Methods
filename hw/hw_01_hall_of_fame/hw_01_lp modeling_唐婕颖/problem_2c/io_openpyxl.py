#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time: 2025/10/6 22:10
# @Author: TJY

"""
Reading from and writing to an Excel file using openpyxl.
"""

from openpyxl import load_workbook

import constant


def readDataOpenpyxl() -> tuple[dict, dict, dict,int]:
    """
    Given an Excel file, read production planning data from Excel file using openpyxl.

    Returns
    -------
    productProfits : dict
        The profit per unit for each product
        - Keys: product names (`Product 1`, `Product 2`)
        - Values: profit per unit
    resourceProductUnits : nested dict
        units of each resource required to produce one unit of each product
        - Keys: resource names (`Metal Frame Parts`, `Electrical Components`)
        - Values: dict
            - Keys: product names (`Product 1`, `Product 2`)
            - Values: units of resource required per unit product
    resourceAvailableUnits : dict
         The available units of each resource
        - Keys: resource names (`Metal Frame Parts`, `Electrical Components`)
        - Values: available units
    product2QuantityLimit : int
        The maximum production quantity limit for Product 2
        - Value: quantity limit
    """

    # Load a workbook from DATA_PATH
    inputBook = load_workbook(constant.DATA_PATH)

    # Find the sheet in which you want to write the solution
    inputSheet = inputBook[constant.INPUT_SHEET_NAME]

    # Read data from the inputSheet and create dictionaries
    productProfits = {}
    for productName in constant.PRODUCT_NAMES:
        colIndex = constant.INPUT_PROFIT_START_COL + constant.PRODUCT_NAMES.index(productName)
        profitValue = inputSheet.cell(constant.INPUT_PROFIT_START_ROW, colIndex).value
        productProfits[productName] = profitValue

    resourceAvailableUnits  = {}
    for i, resourseName in enumerate(constant.Resource_NAMES):
        rowIndex = constant.INPUT_UNITS_AVAILABLE_START_ROW + i
        colIndex = constant.INPUT_UNITS_AVAILABLE_START_COL
        units = inputSheet.cell(rowIndex, colIndex).value
        resourceAvailableUnits[resourseName] = units

    resourceProductUnits= {}
    for i, resourseName in enumerate(constant.Resource_NAMES):
        rowIndex = constant.INPUT_Resourse_UNITS_START_ROW + i
        unitsPerProduct = {}
        for productName in constant.PRODUCT_NAMES:
            colIndex = constant.INPUT_Resourse_UNITS_START_COL + constant.PRODUCT_NAMES.index(productName)
            units = inputSheet.cell(rowIndex, colIndex).value
            unitsPerProduct[productName] = units
        resourceProductUnits[resourseName] = unitsPerProduct

    product2QuantityLimit = inputSheet.cell(
        row=constant.INPUT_PRODUCT2_LIMIT_START_ROW,
        column=constant.INPUT_PRODUCT2_LIMIT_START_COL
    ).value

    inputBook.close()
    return productProfits, resourceProductUnits,  resourceAvailableUnits,product2QuantityLimit


def writeDataOpenpyxl(soln, objVal) -> None:
    """
    Write the solution back to the Excel file using openpyxl.

    Parameters
    ----------
    soln : dict
        The optimal solutions
        - Keys: variable names
        - Values: optimal decision variable values
    objVal : float
        The optimal objective function value
    """
    # Load a workbook from DATA_PATH
    outputBook = load_workbook(constant.DATA_PATH)

    # Create the sheet to write the solution
    outputSheet = outputBook.copy_worksheet(outputBook[constant.INPUT_SHEET_NAME])
    outputSheet.title = constant.OUTPUT_SHEET_NAME


    # Write the optimal solutions to the outputSheet
    for j, product in enumerate(constant.PRODUCT_NAMES):
        rowIndex = constant.OUTPUT_UNITS_PRODUCED_START_ROW
        colIndex = constant.OUTPUT_UNITS_PRODUCED_START_COL + j
        outputSheet.cell(rowIndex, colIndex, soln[product])

    outputSheet.cell(constant.OUTPUT_PROFIT_START_ROW, constant.OUTPUT_PROFIT_START_COL, objVal)

    # Save the workbook
    outputBook.save(constant.DATA_PATH)
    outputBook.close()


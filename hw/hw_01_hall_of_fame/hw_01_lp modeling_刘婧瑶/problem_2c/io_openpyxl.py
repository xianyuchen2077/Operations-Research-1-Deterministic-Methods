"""
Reading from and writing to an Excel file using openpyxl.
"""

from openpyxl import load_workbook

import constant


def readDataOpenpyxl() -> tuple[dict, dict, dict]:
    """
    Given an Excel file, read production planning data from Excel file using openpyxl.

    Returns
    -------
    productProfits : dict
        The profit per batch for each product
        - Keys: product names (`Doors`, `Windows`)
        - Values: profit per batch
    plantProductHours : nested dict
        The hours required to produce one batch of each product at each plant
        - Keys: plant names (`Plant 1`, `Plant 2`, `Plant 3`)
        - Values: dict
            - Keys: product names (`Doors` and `Windows`)
            - Values: hours required
    plantAvailableHours : dict
        The available hours at each plant
        - Keys: plant names (`Plant 1`, `Plant 2`, `Plant 3`)
        - Values: available hours
    """

    # Load a workbook from DATA_PATH
    inputBook = load_workbook(constant.DATA_PATH)

    # Find the sheet in which you want to write the solution
    inputSheet = inputBook[constant.SHEET_NAME]

    # Read data from the inputSheet and create dictionaries
    productProfits = {}
    for productName in constant.PRODUCT_NAMES:
        colIndex = constant.INPUT_PROFIT_START_COL + constant.PRODUCT_NAMES.index(productName)
        profitValue = inputSheet.cell(constant.INPUT_PROFIT_START_ROW, colIndex).value
        productProfits[productName] = profitValue

    plantAvailableHours = {}
    for i, plantName in enumerate(constant.PLANT_NAMES):
        rowIndex = constant.INPUT_HOURS_AVAILABLE_START_ROW + i
        colIndex = constant.INPUT_HOURS_AVAILABLE_START_COL
        hours = inputSheet.cell(rowIndex, colIndex).value
        plantAvailableHours[plantName] = hours

    plantProductHours = {}
    for i, plantName in enumerate(constant.PLANT_NAMES):
        rowIndex = constant.INPUT_HOURS_START_ROW + i
        hoursPerProduct = {}
        for productName in constant.PRODUCT_NAMES:
            colIndex = constant.INPUT_HOURS_START_COL + constant.PRODUCT_NAMES.index(productName)
            hours = inputSheet.cell(rowIndex, colIndex).value
            hoursPerProduct[productName] = hours
        plantProductHours[plantName] = hoursPerProduct

    return productProfits, plantProductHours, plantAvailableHours


def writeDataOpenpyxl(soln, objVal) -> None:
    """
    Write the solution back to the Excel file using openpyxl.

    Parameters
    ----------
    soln : dict
        The optimal solutions
        - Keys: variable names
        - Values: optimal decision variable values
    objVal : float
        The optimal objective function value
    """
    outputBook = load_workbook(constant.DATA_PATH)
    
    # 删除已存在的Solution_Gurobi工作表（如果存在）
    if constant.SOLUTION_SHEET_NAME in outputBook.sheetnames:
        del outputBook[constant.SOLUTION_SHEET_NAME]
    
    # 创建新的解决方案工作表
    solution_sheet = outputBook.create_sheet(constant.SOLUTION_SHEET_NAME)
    
    # 写入列标题（让输出更可读）
    solution_sheet.cell(1, constant.OUTPUT_BATCHES_PRODUCED_START_COL, "Product 1")
    solution_sheet.cell(1, constant.OUTPUT_BATCHES_PRODUCED_START_COL + 1, "Product 2")
    solution_sheet.cell(1, constant.OUTPUT_PROFIT_START_COL, "Total Profit")
    
    # 写入行标签
    solution_sheet.cell(constant.OUTPUT_BATCHES_PRODUCED_START_ROW, 1, "Units Produced")
    
    # 写入最优解
    for j, product in enumerate(constant.PRODUCT_NAMES):
        rowIndex = constant.OUTPUT_BATCHES_PRODUCED_START_ROW
        colIndex = constant.OUTPUT_BATCHES_PRODUCED_START_COL + j
        solution_sheet.cell(rowIndex, colIndex, soln[product])
    
    # 写入总利润
    solution_sheet.cell(constant.OUTPUT_PROFIT_START_ROW, constant.OUTPUT_PROFIT_START_COL, objVal)
    
    # 可选：添加问题描述，让输出更完整
    solution_sheet.cell(4, 1, "Optimal Solution Summary:")
    solution_sheet.cell(5, 1, f"Produce {soln['Product 1']} units of Product 1")
    solution_sheet.cell(6, 1, f"Produce {soln['Product 2']} units of Product 2") 
    solution_sheet.cell(7, 1, f"Maximum Total Profit: ${objVal}")
    
    # Save the workbook
    outputBook.save(constant.DATA_PATH)
    
    print(f"解决方案已写入工作表: {constant.SOLUTION_SHEET_NAME}")
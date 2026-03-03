"""
Reading from and writing to an Excel file using openpyxl.
"""

from openpyxl import load_workbook

import constant
import math


def _to_float_or_none(x):
    """
    Try to convert x to float. Return None if conversion not possible.
    Treat common textual representations of infinity as math.inf.
    """
    if x is None:
        return None
    if isinstance(x, (int, float)):
        # guard against NaN
        try:
            f = float(x)
        except Exception:
            return None
        if math.isnan(f):
            return None
        return f

    try:
        s = str(x).strip()
    except Exception:
        return None
    if not s:
        return None
    low = s.lower()
    if low in ("inf", "infinity", "unlimited", "无限"):
        return math.inf
    try:
        return float(s)
    except Exception:
        return None


def readDataOpenpyxl() -> tuple[dict, dict, dict, dict]:
    """
    Given an Excel file, read production planning data from Excel file using openpyxl.

    Returns
    -------
    productProfits : dict
        The profit per unit for each product
        - Keys: product names (`Product1`, `Product2`)
        - Values: profit per unit
    materialsPerUnit : nested dict
        The materials required to produce one unit of each product
        - Keys: material names (`Frame parts`, `Electrical components`)
        - Values: dict
            - Keys: product names (`Product1`, `Product2`)
            - Values: materials required
    materialLimitations : dict
        The material limitations
        - Keys: material names (`Frame parts`, `Electrical components`)
        - Values: available materials
    numberLimitations : dict
        The number limitations
        - Keys: product names (`Product1`, `Product2` and `max`, `min`)
        - Values: available products
    """

    # Load a workbook from DATA_PATH
    inputBook = load_workbook(constant.DATA_PATH)

    # Find the sheet in which you want to read the data
    inputSheet = inputBook[constant.SHEET_NAME_READ]

    # Read data from the inputSheet and create dictionaries
    # Read profit per unit
    productProfits = {}
    for productName in constant.PRODUCT_NAMES:
        colIndex = constant.INPUT_PROFIT_START_COL + constant.PRODUCT_NAMES.index(productName)
        profitValue = inputSheet.cell(constant.INPUT_PROFIT_ROW, colIndex).value
        productProfits[productName] = profitValue

    # Read materials used per unit produced
    materialsPerUnit = {}
    for i, materialName in enumerate(constant.MATERIAL_NAMES):
        rowIndex = constant.INPUT_MATERIALS_START_ROW + i
        materialsPerProduct = {}
        for j, productName in enumerate(constant.PRODUCT_NAMES):
            colIndex = constant.INPUT_MATERIALS_START_COL + j
            materials = inputSheet.cell(rowIndex, colIndex).value
            materialsPerProduct[productName] = materials
        materialsPerUnit[materialName] = materialsPerProduct

    # Read material limitations
    materialLimitations = {}
    for i, materialName in enumerate(constant.MATERIAL_NAMES):
        rowIndex = constant.INPUT_MATERIAL_LIMITATIONS_START_ROW + i
        colIndex = constant.INPUT_MATERIAL_LIMITATIONS_COL
        limitation = inputSheet.cell(rowIndex, colIndex).value
        materialLimitations[materialName] = limitation

    # Read number limitations
    numberLimitations = {}
    for i in range(constant.INPUT_NUMBER_LIMITATIONS_START_ROW,constant.INPUT_NUMBER_LIMITATIONS_END_ROW + 1):
        productName = None
        rowIndex = i
        colIndex = constant.INPUT_NUMBER_LIMITATIONS_NAME_COL
        if inputSheet.cell(rowIndex, colIndex).value == "Number of product1":
            productName = "Product1"
        elif inputSheet.cell(rowIndex, colIndex).value == "Number of product2":
            productName = "Product2"

        if productName:
            productLimitations = numberLimitations.get(productName)
            if productLimitations is None:
                productLimitations = {}
                numberLimitations[productName] = productLimitations

            colIndex = constant.INPUT_NUMBER_LIMITATIONS_OPERATOR_COL

            # Ensure whether it is upper bound or lower bound
            operator = inputSheet.cell(rowIndex, colIndex).value
            if operator == "<=":
                op_key = "max"
            elif operator == ">=":
                op_key = "min"
            else:
                op_key = None

            # Ensure multiple limitations for the same product (min and max) are merged
            if op_key:
                colIndex = constant.INPUT_NUMBER_LIMITATIONS_COL
                limitation_raw = inputSheet.cell(rowIndex, colIndex).value
                limitation_value = _to_float_or_none(limitation_raw)

                if op_key == "max":
                    if limitation_value is None:
                        limitation_value = math.inf
                    prev_raw = productLimitations.get('max')
                    prev = _to_float_or_none(prev_raw)
                    if prev is not None:
                        limitation_value = min(limitation_value, prev)
                    productLimitations['max'] = limitation_value
                elif op_key == "min":
                    if limitation_value is None:
                        limitation_value = 0.0
                    prev_raw = productLimitations.get('min')
                    prev = _to_float_or_none(prev_raw)
                    if prev is not None:
                        limitation_value = max(prev, limitation_value)
                    productLimitations['min'] = limitation_value

    return productProfits, materialsPerUnit, materialLimitations, numberLimitations


def writeDataOpenpyxl(soln, objVal, materialsUsed, profit) -> None:
    """
    Write the solution back to the Excel file using openpyxl.

    Parameters
    ----------
    soln : dict
        The optimal solutions
        - Keys: product names
        - Values: optimal number of products
    objVal : float
        The optimal objective function value (total profit)
    materialsUsed : dict
        The total materials used
        - Keys: material names
        - Values: total materials used
    """
    # Load a workbook from DATA_PATH
    outputBook = load_workbook(constant.DATA_PATH)

    # Find the sheet to write the solution
    # if the sheet does not exist, create it
    if constant.SHEET_NAME_WRITE not in outputBook.sheetnames:
        outputBook.create_sheet(constant.SHEET_NAME_WRITE)
    outputSheet = outputBook[constant.SHEET_NAME_WRITE]

    # Write basic information(title, description, etc.)
    outputSheet.cell(constant.OUTPUT_PRODUCTS_ROW ,constant.OUTPUT_PRODUCTS_START_COL - 1, 'Number of Products')
    outputSheet.cell(constant.OUTPUT_PRODUCTS_ROW - 1 ,constant.OUTPUT_TOTAL_PROFIT_COL, 'Total')
    outputSheet.cell(constant.OUTPUT_TOTAL_PROFIT_ROW ,constant.OUTPUT_PRODUCTS_START_COL - 1, 'Profit')
    # Write the optimal number of products
    for j, product in enumerate(constant.PRODUCT_NAMES):
        rowIndex = constant.OUTPUT_PRODUCTS_ROW
        colIndex = constant.OUTPUT_PRODUCTS_START_COL + j
        outputSheet.cell(rowIndex - 1, colIndex, product)
        outputSheet.cell(rowIndex, colIndex, soln[product])

        if materialsUsed:
            for k, material in enumerate(constant.MATERIAL_NAMES):
                rowIndex = constant.OUTPUT_PRODUCTS_ROW + k + 1
                outputSheet.cell(rowIndex, colIndex, materialsUsed[material][product])

        outputSheet.cell(constant.OUTPUT_TOTAL_PROFIT_ROW, colIndex, profit[product])

    # Write the total profit
    outputSheet.cell(constant.OUTPUT_TOTAL_PROFIT_ROW, constant.OUTPUT_TOTAL_PROFIT_COL, objVal)

    # Write the materials used in total
    if materialsUsed:
        for i, materialName in enumerate(constant.MATERIAL_NAMES):
            rowIndex = constant.OUTPUT_MATERIALS_TOTAL_START_ROW + i
            colIndex = constant.OUTPUT_MATERIALS_TOTAL_COL
            outputSheet.cell(rowIndex, constant.OUTPUT_PRODUCTS_START_COL - 1, f'{materialName}')
            outputSheet.cell(rowIndex, constant.OUTPUT_TOTAL_PROFIT_COL, materialsUsed[materialName]["Total"])

    # Center and adjust column widths
    autosize_columns(outputSheet, min_width=8.0, padding=2.0)
    center_range(outputSheet, cell_range=None, wrap_text=True)

    # Save the workbook
    outputBook.save(constant.DATA_PATH)

from openpyxl.utils import get_column_letter
def autosize_columns(ws, min_width: float = 8.0, padding: float = 2.0):
    """
    adjust column widths according to the max text length in each column
    ws: openpyxl worksheet
    min_width: Minimum column width each character
    padding: Additional width to add to the maximum length
    """
    for col_cells in ws.columns:
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            # Convert the value to a string and calculate its length (ignoring line breaks)
            s = str(cell.value)
            s_line = max(s.splitlines(), key=len) if '\n' in s else s
            max_len = max(max_len, len(s_line))
        adjusted_width = max(min_width, max_len + padding)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = adjusted_width

from openpyxl.styles import Alignment
def center_range(ws, cell_range=None, horizontal='center', vertical='center', wrap_text=False):
    """
    center all text in the specified range or the entire sheet
    ws: worksheet
    cell_range: e.g. "A1:D20" or None(all sheet)
    horizontal: 'center' | 'left' | 'right'
    vertical: 'center' | 'top' | 'bottom'
    wrap_text: Whether to wrap text
    """
    align = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)
    if cell_range is None:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = align
    else:
        min_col, min_row, max_col, max_row = ws[cell_range].min_row, ws[cell_range].min_col, ws[cell_range].max_col, ws[cell_range].max_row
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.alignment = align
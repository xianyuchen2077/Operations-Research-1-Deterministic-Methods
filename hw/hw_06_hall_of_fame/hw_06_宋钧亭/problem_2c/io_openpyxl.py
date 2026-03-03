"""
Reading from and writing to an Excel file using openpyxl.
"""

from openpyxl import load_workbook

import constant


def readDataOpenpyxl() -> tuple[dict, dict, dict, dict, int]:
    """Read data from an Excel file using openpyxl.
    
    Returns
    -------
    computerCost: dict
        The cost per computer for each vendor.
        - Keys: vendor index
        - Values: computer cost
    deliveryCost: dict
        The cost of delivery for each vendor.
        - Keys: vendor index
        - Values: delivery cost
    maxQuantity: dict
        The maximum quantity to be ordered for each vendor.
        - Keys: vendor index
        - Values: maximum quantity
    minQuantity: dict
        The minimum quantity to be ordered for each vendor.
        - Keys: vendor index
        - Values: minimum quantity
    totalDemand: int
        The total demand for state university
    """

    # Load a workbook from DATA_PATH
    inputBook = load_workbook(constant.DATA_PATH)

    # Find the sheet in which you want to read data
    inputSheet = inputBook[constant.INPUT_SHEET_NAME]

    # Read data from the inputSheet and create dictionaries
    computerCost = {}
    for i, vendorName in enumerate(constant.VENDOR_NAMES):
        colIndex = constant.INPUT_COMPUTER_COST_START_COL + i
        costValue = inputSheet.cell(constant.INPUT_COMPUTER_COST_START_ROW, colIndex).value
        computerCost[vendorName] = costValue

    deliveryCost = {}
    for i, vendorName in enumerate(constant.VENDOR_NAMES):
        colIndex = constant.INPUT_DELIVERY_COST_START_COL + i
        costValue = inputSheet.cell(constant.INPUT_DELIVERY_COST_START_ROW, colIndex).value
        deliveryCost[vendorName] = costValue

    maxQuantity = {}
    for i, vendorName in enumerate(constant.VENDOR_NAMES):
        colIndex = constant.INPUT_MAX_QUANTITY_START_COL + i
        costValue = inputSheet.cell(constant.INPUT_MAX_QUANTITY_START_ROW, colIndex).value
        maxQuantity[vendorName] = costValue

    minQuantity = {}
    for i, vendorName in enumerate(constant.VENDOR_NAMES):
        colIndex = constant.INPUT_MIN_QUANTITY_START_COL + i
        costValue = inputSheet.cell(constant.INPUT_MIN_QUANTITY_START_ROW, colIndex).value
        minQuantity[vendorName] = costValue

    totalDemand = inputSheet.cell(constant.INPUT_TOTAL_DEMAND_START_ROW,
                                  constant.INPUT_TOTAL_DEMAND_START_COL).value

    return computerCost, deliveryCost, maxQuantity, minQuantity, totalDemand






def writeDataOpenpyxl(soln, objVal) -> None:
    """
    Write the solution back to the Excel file using openpyxl.
    
    Parameters
    ----------
    soln : dict
        The selection of ordering or not and the quantity of ordering
        - Keys: whether to order[investment index] or quantity to order[investment index]
        - Values:  (0 or 1) or int
    objVal : int
        The total cost.
    """

    # Load a workbook from DATA_PATH
    outputBook = load_workbook(constant.DATA_PATH)

    # Find the sheet to write the solution
    if constant.OUTPUT_SHEET_NAME in outputBook.sheetnames:
        del outputBook[constant.OUTPUT_SHEET_NAME]
    outputSheet = outputBook.create_sheet(constant.OUTPUT_SHEET_NAME)

    # head
    outputSheet.cell(constant.OUTPUT_WHETHER_TO_ORDER_NAME_ROW,constant.OUTPUT_WHETHER_TO_ORDER_NAME_COL).value \
        = constant.WHETHER_TO_ORDER_NAME
    outputSheet.cell(constant.OUTPUT_QUANTITY_TO_ORDER_NAME_ROW,constant.OUTPUT_QUANTITY_TO_ORDER_NAME_COL).value \
        = constant.QUANTITY_TO_ORDER_NAME
    outputSheet.cell(constant.OUTPUT_TOTAL_COST_ROW,constant.OUTPUT_TOTAL_COST_COL).value \
        = constant.TOTAL_COST_NAME
    for i, vendorName in enumerate(constant.VENDOR_NAMES):
        col = constant.OUTPUT_VENDOR_NAME_START_COL + i
        outputSheet.cell(constant.OUTPUT_VENDOR_NAME_START_ROW, col).value = vendorName

    # Write data to the outputSheet
    for i, vendorName in enumerate(constant.VENDOR_NAMES):
        colIndex = constant.OUTPUT_WHETHER_TO_ORDER_START_COL + i
        varName = f"{constant.WHETHER_TO_ORDER_NAME}[{vendorName}]"
        outputSheet.cell(constant.OUTPUT_WHETHER_TO_ORDER_START_ROW, colIndex).value = soln[varName]

    for i, vendorName in enumerate(constant.VENDOR_NAMES):
        colIndex = constant.OUTPUT_QUANTITY_TO_ORDER_START_COL + i
        varName = f"{constant.QUANTITY_TO_ORDER_NAME}[{vendorName}]"
        outputSheet.cell(constant.OUTPUT_QUANTITY_TO_ORDER_START_ROW, colIndex).value = soln[varName]

    outputSheet.cell(constant.OUTPUT_MINIMUM_COST_START_ROW, constant.OUTPUT_MINIMUM_COST_START_COL).value = objVal

    # Save the workbook to DATA_PATH
    outputBook.save(constant.DATA_PATH)

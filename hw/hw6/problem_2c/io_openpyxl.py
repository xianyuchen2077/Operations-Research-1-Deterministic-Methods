"""
Reading from and writing to an Excel file using openpyxl.
"""

from openpyxl import load_workbook

import constant


def readDataOpenpyxl() -> tuple[dict, dict, dict, dict, dict]:
    """Read data from an Excel file using openpyxl.

    Returns
    -------
    unit_Price : dict
        The unit price for each vendor.
        - Keys: vendor name
        - Values: unit price
    delivery_Charge : dict
        The delivery charge for each vendor.
        - Keys: vendor name
        - Values: delivery charge
    max_purchase_quantity : dict
        The maximum quantity each vendor can supply.
        - Keys: vendor name
        - Values: maximum quantity
    min_purchase_quantity : dict
        The minimum required purchase quantity for each vendor.
        - Keys: vendor name
        - Values: minimum quantity
    planned_computer_Number : dict
        The total_number_of_computers_planned_to_purchase.
        - Key: "total_number_of_computers_planned_to_purchase"
        - Values: planned number
    """

    # Load a workbook from DATA_PATH
    inputBook = load_workbook(constant.DATA_PATH)

    # Find the sheet in which you want to read data
    inputSheet = inputBook[constant.SHEET_NAME_INPUT]

    # Read data from the inputSheet and create dictionaries
    unit_Price = {}
    for i, vendorName in enumerate(constant.VENDOR_NAMES):
        colIndex = constant.UNIT_PRICE_START_COL + i
        costValue = inputSheet.cell(constant.UNIT_PRICE_START_ROW, colIndex).value
        unit_Price[vendorName] = costValue

    delivery_Charge = {}
    for i, vendorName in enumerate(constant.VENDOR_NAMES):
        colIndex = constant.DELIVERY_CHARGE_START_COL + i
        value = inputSheet.cell(constant.DELIVERY_CHARGE_START_ROW, colIndex).value
        delivery_Charge[vendorName] = value

    max_purchase_quantity = {}
    for i, vendorName in enumerate(constant.VENDOR_NAMES):
        colIndex = constant.MAX_QUANTITY_START_COL + i
        value = inputSheet.cell(constant.MAX_QUANTITY_START_ROW, colIndex).value
        max_purchase_quantity[vendorName] = value

    min_purchase_quantity = {}
    for i, vendorName in enumerate(constant.VENDOR_NAMES):
        colIndex = constant.MIN_PURCHASE_QUANTITY_START_COL + i
        value = inputSheet.cell(constant.MIN_PURCHASE_QUANTITY_START_ROW, colIndex).value
        min_purchase_quantity[vendorName] = value

    planned_computer_Number = {}
    planned_computer_Number[constant.PLANNED_COMPUTER_NUMBER] = inputSheet.cell(constant.PLANNED_COMPUTER_NUMBER_START_ROW, constant.PLANNED_COMPUTER_NUMBER_START_COL).value

    return unit_Price, delivery_Charge, max_purchase_quantity, min_purchase_quantity, planned_computer_Number

def writeDataOpenpyxl(soln, objVal) -> None:
    """
    Write the solution back to the Excel file using openpyxl.

    Parameters
    ----------
    soln : dict
        The purchase numbers for each vendor.
        - Keys: vendor name
        - Values: purchase number
    objVal : float
        The total NPV achieved.
    """

    # Load a workbook from DATA_PATH
    outputBook = load_workbook(constant.DATA_PATH)

    # Find the sheet to write the solution
    outputSheet = outputBook[constant.SHEET_NAME_OUTPUT]

    # Write data to the outputSheet
    for i, vendorName in enumerate(soln.keys()):
        if (i <= 2):
            colIndex = constant.OUTPUT_WHETHER_PURCHASE_START_COL + i
            outputSheet.cell(constant.OUTPUT_WHETHER_PURCHASE_START_ROW, colIndex).value = soln[vendorName]
        else:
            colIndex = constant.OUTPUT_PURCHASE_NUMBERS_START_COL + i - 3
            outputSheet.cell(constant.OUTPUT_PURCHASE_NUMBERS_START_ROW, colIndex).value = soln[vendorName]

    outputSheet.cell(constant.OUTPUT_MIN_TOTAL_COST_START_ROW, constant.OUTPUT_MIN_TOTAL_COST_START_COL).value = objVal

    for i, vendorName in enumerate(constant.VENDOR_NAMES):
        colIndex = constant.OUTPUT_TITLE_COL + i
        outputSheet.cell(constant.OUTPUT_TITLE_ROW, colIndex,'').value = vendorName
    outputSheet.cell(constant.OUTPUT_WHETHER_PURCHASE_START_ROW, constant.OUTPUT_TITLE_COL - 1,'').value = "Purchase any？"
    outputSheet.cell(constant.OUTPUT_PURCHASE_NUMBERS_START_ROW, constant.OUTPUT_TITLE_COL - 1,'').value = "Purchase Numbers"
    outputSheet.cell(constant.OUTPUT_MIN_TOTAL_COST_START_ROW, constant.OUTPUT_TITLE_COL - 1,'').value = "Minimum Total Cost"

    # Center and adjust column widths
    autosize_columns(outputSheet, min_width=8.0, padding=2.0)
    center_range(outputSheet, cell_range=None, wrap_text=True)

    # Save the workbook to DATA_PATH
    outputBook.save(constant.DATA_PATH)
    outputBook.save(constant.DATA_PATH)

# extra utility functions for formatting
from openpyxl.utils import get_column_letter
def autosize_columns(ws, min_width: float = 8.0, padding: float = 2.0):
    """
    adjust column widths according to the max text length in each column
    ws: openpyxl worksheet
    min_width: Minimum column width each character
    padding: Additional width to add to the maximum length
    """
    for col_cells in ws.columns:
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            # Convert the value to a string and calculate its length (ignoring line breaks)
            s = str(cell.value)
            s_line = max(s.splitlines(), key=len) if '\n' in s else s
            max_len = max(max_len, len(s_line))
        adjusted_width = max(min_width, max_len + padding)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = adjusted_width

from openpyxl.styles import Alignment
def center_range(ws, cell_range=None, horizontal='center', vertical='center', wrap_text=False):
    """
    center all text in the specified range or the entire sheet
    ws: worksheet
    cell_range: e.g. "A1:D20" or None(all sheet)
    horizontal: 'center' | 'left' | 'right'
    vertical: 'center' | 'top' | 'bottom'
    wrap_text: Whether to wrap text
    """
    align = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)
    if cell_range is None:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = align
    else:
        min_col, min_row, max_col, max_row = ws[cell_range].min_row, ws[cell_range].min_col, ws[cell_range].max_col, ws[cell_range].max_row
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.alignment = align
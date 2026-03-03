"""
Reading from and writing to an Excel file using openpyxl.
"""

from openpyxl import load_workbook

import constant


def readDataOpenpyxl() -> tuple[dict, dict, dict]:
    """Read data from an Excel file using openpyxl.
    
    Returns
    -------
    investmentCost : dict
        The investment cost for each investment.
        - Keys: investment index
        - Values: investment cost
    npv : dict
        The NPV for each investment.
        - Keys: investment index
        - Values: NPV
    budget : dict
        The budget available for investments.
        - Key: "Budget"
        - Values: budget amount
    """

    # Load a workbook from DATA_PATH
    inputBook = load_workbook(constant.DATA_PATH)

    # Find the sheet in which you want to read data
    inputSheet = inputBook[constant.SHEET_NAME]

    # Read data from the inputSheet and create dictionaries
    investmentCost = {}
    for i, investmentName in enumerate(constant.INVESTMENT_NAMES):
        colIndex = constant.INPUT_INVESTMENT_COST_START_COL + i
        costValue = inputSheet.cell(constant.INPUT_INVESTMENT_COST_START_ROW, colIndex).value
        investmentCost[investmentName] = costValue

    npv = {}
    for i, investmentName in enumerate(constant.INVESTMENT_NAMES):
        colIndex = constant.INPUT_NPV_START_COL + i
        value = inputSheet.cell(constant.INPUT_NPV_START_ROW, colIndex).value
        npv[investmentName] = value
    
    budget = {}
    budget[constant.BUDGET_NAME] = inputSheet.cell(constant.INPUT_BUDGET_START_ROW, constant.INPUT_BUDGET_START_COL).value
    
    return investmentCost, npv, budget


def writeDataOpenpyxl(soln, objVal) -> None:
    """
    Write the solution back to the Excel file using openpyxl.
    
    Parameters
    ----------
    soln : dict
        The investment levels for each investment.
        - Keys: investment index
        - Values: investment level (0 or 1)
    objVal : float
        The total NPV achieved.
    """

    # Load a workbook from DATA_PATH
    outputBook = load_workbook(constant.DATA_PATH)

    # Find the sheet to write the solution
    outputSheet = outputBook[constant.SHEET_NAME]

    # Write data to the outputSheet
    for i, investmentName in enumerate(constant.INVESTMENT_NAMES):
        colIndex = constant.OUTPUT_INVESTMENT_LEVELS_START_COL + i
        outputSheet.cell(constant.OUTPUT_INVESTMENT_LEVELS_START_ROW, colIndex).value = soln[investmentName]

    outputSheet.cell(constant.OUTPUT_TOTAL_NPV_START_ROW, constant.OUTPUT_TOTAL_NPV_START_COL).value = objVal

    # Save the workbook to DATA_PATH
    outputBook.save(constant.DATA_PATH)
